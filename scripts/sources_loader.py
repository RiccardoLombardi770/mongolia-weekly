"""
sources_loader.py — read sources/sources.xlsx (the office's shared workbook).

Design rule: a bad row never stops the Monday run. Every problem is collected
into `issues`, which the workflow prints into the pull request under "Source
problems" so it can be fixed in the spreadsheet, not in the code.

Falls back to config.yaml if the workbook is missing, so the system still runs
on a fresh clone.

To move the workbook to a Google Sheet later, fill workbook.csv_urls in
config.yaml (File > Share > Publish to web > CSV, one URL per sheet). Nothing
else changes.
"""

import csv
import io
import pathlib

import yaml

ROOT = pathlib.Path(__file__).resolve().parent.parent
CONFIG = yaml.safe_load((ROOT / "config.yaml").read_text(encoding="utf-8"))

EXPECTED = {
    "sources": ["name", "type", "query_or_url", "language", "default_category",
                "active", "notes"],
    "outlets": ["name", "domain", "language", "country", "priority", "rss_url",
                "active", "notes"],
    "roster": ["name", "role", "agency", "aliases", "active", "notes"],
    "un_sites": ["agency", "name", "url", "feed_url", "active", "notes"],
    "glossary": ["english", "mongolian", "notes"],
}


def _truthy(v):
    return str(v or "").strip().lower() in ("yes", "y", "true", "1", "si", "sì", "тийм")


def _clean(v):
    return str(v).strip() if v is not None else ""


def _is_placeholder(row):
    joined = " ".join(_clean(v) for v in row.values()).upper()
    return "EXAMPLE — DELETE" in joined or "EXAMPLE - DELETE" in joined or "<FILL IN>" in joined


def _rows_from_sheet(ws, sheet, issues):
    """Header is row 2 (row 1 is the yellow help banner)."""
    grid = list(ws.iter_rows(values_only=True))
    if len(grid) < 2:
        issues.append(f"[{sheet}] sheet is empty")
        return []
    header = [_clean(h).lower() for h in grid[1]]
    missing = [c for c in EXPECTED.get(sheet, []) if c not in header]
    if missing:
        issues.append(f"[{sheet}] missing column(s): {', '.join(missing)} — "
                      f"header names must not be changed")
    out = []
    for n, raw in enumerate(grid[2:], start=3):
        row = {header[i]: _clean(v) for i, v in enumerate(raw) if i < len(header) and header[i]}
        if not any(row.values()):
            continue
        row["_row"] = n
        row["_sheet"] = sheet
        out.append(row)
    return out


def _rows_from_csv(text, sheet, issues):
    reader = csv.reader(io.StringIO(text))
    grid = [r for r in reader]
    if len(grid) < 2:
        issues.append(f"[{sheet}] published CSV is empty")
        return []
    # Published Google Sheets keep the same layout: banner row, then header.
    header_idx = 1 if len(grid) > 1 and any(
        c.strip().lower() in EXPECTED.get(sheet, []) for c in grid[1]) else 0
    header = [c.strip().lower() for c in grid[header_idx]]
    out = []
    for n, raw in enumerate(grid[header_idx + 1:], start=header_idx + 2):
        row = {header[i]: c.strip() for i, c in enumerate(raw) if i < len(header) and header[i]}
        if not any(row.values()):
            continue
        row["_row"] = n
        row["_sheet"] = sheet
        out.append(row)
    return out


def load_workbook_data():
    """Return (data, issues). data maps sheet name -> list of active rows."""
    issues = []
    data = {k: [] for k in EXPECTED}

    csv_urls = (CONFIG.get("workbook") or {}).get("csv_urls") or {}
    if csv_urls:
        import requests
        for sheet, url in csv_urls.items():
            try:
                r = requests.get(url, timeout=30)
                r.raise_for_status()
                data[sheet] = _rows_from_csv(r.text, sheet, issues)
            except Exception as e:
                issues.append(f"[{sheet}] could not read the published CSV: {e}")
    else:
        path = ROOT / (CONFIG.get("workbook") or {}).get("path", "sources/sources.xlsx")
        if not path.exists():
            issues.append(f"workbook not found at {path.name} — falling back to "
                          f"the source list in config.yaml")
            return data, issues
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            issues.append(f"workbook could not be opened ({e}) — falling back to config.yaml")
            return data, issues
        for sheet in EXPECTED:
            if sheet not in wb.sheetnames:
                issues.append(f"[{sheet}] sheet is missing from the workbook")
                continue
            data[sheet] = _rows_from_sheet(wb[sheet], sheet, issues)

    # Filter: drop example/placeholder rows and anything set to active = no.
    for sheet, rows in data.items():
        keep = []
        for row in rows:
            if _is_placeholder(row):
                continue
            if "active" in row and row.get("active") and not _truthy(row["active"]):
                continue
            keep.append(row)
        data[sheet] = keep

    _validate(data, issues)
    return data, issues


def _validate(data, issues):
    def where(r):
        return f"[{r['_sheet']} row {r['_row']}]"

    for r in data["sources"]:
        t = r.get("type", "").lower()
        if t not in ("google_news", "rss"):
            issues.append(f"{where(r)} type '{r.get('type')}' is not google_news or rss — row skipped")
        if not r.get("query_or_url"):
            issues.append(f"{where(r)} query_or_url is empty — row skipped")
        if t == "rss" and not r.get("query_or_url", "").startswith("http"):
            issues.append(f"{where(r)} rss row needs a full URL starting with http")
        cat = r.get("default_category")
        if cat and cat not in CONFIG["categories"]:
            issues.append(f"{where(r)} default_category '{cat}' is not one of the four "
                          f"categories — it will be ignored")

    for r in data["outlets"]:
        if not r.get("domain"):
            issues.append(f"{where(r)} outlet has no domain — row skipped")
        elif "/" in r["domain"] or r["domain"].startswith("http"):
            issues.append(f"{where(r)} domain should be just the host, e.g. news.mn")

    for r in data["roster"]:
        if not r.get("name"):
            issues.append(f"{where(r)} roster entry has no name — row skipped")
    if not [r for r in data["roster"] if r.get("name") and not r["name"].startswith("<")]:
        issues.append("[roster] no names filled in yet — mentions that quote the "
                      "Resident Coordinator or a head of office by name cannot be "
                      "detected until the sheet is completed")

    for r in data["un_sites"]:
        if not r.get("url", "").startswith("http"):
            issues.append(f"{where(r)} un_sites url must start with http — row skipped")


# ---------------------------------------------------------------- accessors

def news_sources(data):
    """Return (google_news_queries, rss_feeds) for collect.py."""
    queries, feeds = [], []
    for r in data["sources"]:
        t, v = r.get("type", "").lower(), r.get("query_or_url", "")
        if not v:
            continue
        if t == "google_news":
            queries.append({"query": v, "language": r.get("language", "en"),
                            "category": r.get("default_category", "")})
        elif t == "rss" and v.startswith("http"):
            feeds.append({"name": r.get("name") or v, "url": v,
                          "language": r.get("language", "en"),
                          "category": r.get("default_category", "")})
    if not queries and not feeds:
        cfg = CONFIG["sources"]
        queries = [{"query": q, "language": "en", "category": ""}
                   for q in cfg.get("google_news", {}).get("queries", [])]
        feeds = [{"name": f["name"], "url": f["url"], "language": "en", "category": ""}
                 for f in cfg.get("rss_feeds", [])]
    return queries, feeds


def outlets(data):
    out = []
    for r in data["outlets"]:
        if not r.get("domain"):
            continue
        try:
            priority = int(float(r.get("priority") or 2))
        except ValueError:
            priority = 2
        out.append({"name": r.get("name") or r["domain"],
                    "domain": r["domain"].lower().replace("www.", ""),
                    "language": r.get("language", "en"),
                    "country": r.get("country", ""),
                    "priority": priority,
                    "rss_url": r.get("rss_url", "")})
    return out


def roster(data):
    out = []
    for r in data["roster"]:
        name = r.get("name", "")
        if not name or name.startswith("<"):
            continue
        aliases = [a.strip() for a in (r.get("aliases") or "").split(",") if a.strip()]
        out.append({"name": name, "role": r.get("role", ""),
                    "agency": r.get("agency", ""), "aliases": aliases})
    return out


def un_sites(data):
    out = [{"agency": r.get("agency", ""), "name": r.get("name", ""),
            "url": r["url"], "feed_url": r.get("feed_url", "")}
           for r in data["un_sites"] if r.get("url", "").startswith("http")]
    return out


def glossary(data):
    return {r["english"]: r["mongolian"] for r in data["glossary"]
            if r.get("english") and r.get("mongolian")}


def protected_names(data):
    """Outlet names must never be rewritten by the style linter."""
    return [r["name"] for r in data["outlets"] if r.get("name")]


if __name__ == "__main__":
    d, iss = load_workbook_data()
    for k, v in d.items():
        print(f"{k:10s} {len(v)} active row(s)")
    print(f"\n{len(iss)} issue(s):")
    for i in iss:
        print("  -", i)
