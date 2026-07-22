"""
make_sources_xlsx.py — create the shared workbook the office edits.

Run this ONCE to produce sources/sources.xlsx. After that the file belongs to
the office: they open it in Excel, add rows, and upload it back to GitHub.
Re-running overwrites their edits, so it refuses unless you pass --force.

    python scripts/make_sources_xlsx.py
    python scripts/make_sources_xlsx.py --force
"""

import sys
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "sources" / "sources.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="1C1A17")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
EXAMPLE_FONT = Font(name="Arial", size=10, italic=True, color="808080")
NOTE_FILL = PatternFill("solid", fgColor="FFF7DC")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

SHEETS = {
    # ---------------------------------------------------------------- sources
    "sources": {
        "help": "Feeds and searches used to build the weekly news report. "
                "type = google_news (put a search phrase in query_or_url) or "
                "rss (put the full feed URL).",
        "cols": ["name", "type", "query_or_url", "language", "default_category",
                 "active", "notes"],
        "example": ["EXAMPLE — delete this row", "google_news",
                    "Mongolia herders winter", "en", "Social and Humanitarian",
                    "yes", "How a row should look"],
        "rows": [
            ["Google News — economy", "google_news", "Mongolia economy", "en",
             "Economy and Development", "yes", ""],
            ["Google News — mining and coal", "google_news",
             "Mongolia mining coal exports", "en", "Economy and Development",
             "yes", ""],
            ["Google News — politics", "google_news",
             "Mongolia politics government parliament", "en",
             "Politics, Governance and Diplomacy", "yes", ""],
            ["Google News — diplomacy", "google_news",
             "Mongolia China Russia diplomacy", "en",
             "Politics, Governance and Diplomacy", "yes", ""],
            ["Google News — climate", "google_news",
             "Mongolia climate desertification air pollution", "en",
             "Environment and Climate Change", "yes", ""],
            ["Google News — dzud and herders", "google_news",
             "Mongolia dzud winter herders", "en", "Social and Humanitarian",
             "yes", ""],
            ["Google News — health and social", "google_news",
             "Mongolia health education social protection", "en",
             "Social and Humanitarian", "yes", ""],
            ["Google News — Ulaanbaatar", "google_news", "Ulaanbaatar", "en", "",
             "yes", ""],
            ["Google News — Mongolian language", "google_news", "Монгол улс",
             "mn", "", "yes", "hl/gl switch to mn-MN automatically"],
            ["UN News", "rss",
             "https://news.un.org/feed/subscribe/en/news/all/rss.xml", "en", "",
             "yes", ""],
            ["MONTSAME (state news agency)", "rss",
             "https://montsame.mn/en/rss", "en", "", "no",
             "VERIFY the feed URL before switching to yes"],
            ["News.mn", "rss", "https://news.mn/feed/", "mn", "", "no",
             "VERIFY the feed URL before switching to yes"],
        ],
    },
    # ---------------------------------------------------------------- outlets
    "outlets": {
        "help": "Media outlets to watch for mentions of the United Nations in "
                "Mongolia. priority 1 = national/most important, 3 = minor. "
                "rss_url is optional; leave it blank and the outlet is still "
                "picked up through search.",
        "cols": ["name", "domain", "language", "country", "priority", "rss_url",
                 "active", "notes"],
        "example": ["EXAMPLE — delete this row", "example.mn", "mn", "Mongolia",
                    "1", "", "yes", "How a row should look"],
        "rows": [
            ["MONTSAME", "montsame.mn", "mn", "Mongolia", 1, "", "yes",
             "State news agency"],
            ["News.mn", "news.mn", "mn", "Mongolia", 1, "", "yes", ""],
            ["iKon.mn", "ikon.mn", "mn", "Mongolia", 1, "", "yes", ""],
            ["Gogo.mn", "gogo.mn", "mn", "Mongolia", 1, "", "yes", ""],
            ["Zindaa.mn", "zindaa.mn", "mn", "Mongolia", 2, "", "yes", ""],
            ["Unuudur (Өдрийн сонин)", "unuudur.mn", "mn", "Mongolia", 2, "",
             "yes", ""],
            ["Eguur.mn", "eguur.mn", "mn", "Mongolia", 3, "", "yes", ""],
            ["Isee.mn", "isee.mn", "mn", "Mongolia", 3, "", "yes", ""],
            ["Polit.mn", "polit.mn", "mn", "Mongolia", 3, "", "yes", ""],
            ["The UB Post", "ubposts.mn", "en", "Mongolia", 1, "", "yes",
             "VERIFY current domain"],
            ["Mongolian National Broadcaster", "mnb.mn", "mn", "Mongolia", 2, "",
             "yes", ""],
            ["Bloomberg TV Mongolia", "bloombergtv.mn", "mn", "Mongolia", 2, "",
             "yes", ""],
            ["Reuters", "reuters.com", "en", "International", 1, "", "yes", ""],
            ["Xinhua", "news.cn", "en", "International", 2, "", "yes", ""],
            ["TASS", "tass.com", "en", "International", 2, "", "yes", ""],
            ["AKIpress", "akipress.com", "en", "International", 2, "", "yes", ""],
            ["The Diplomat", "thediplomat.com", "en", "International", 2, "",
             "yes", ""],
            ["Nikkei Asia", "asia.nikkei.com", "en", "International", 3, "",
             "yes", ""],
        ],
    },
    # ----------------------------------------------------------------- roster
    "roster": {
        "help": "Names the media may quote. aliases = every other spelling you "
                "have seen, separated by commas — Cyrillic, inverted name "
                "order, common misspellings. The more aliases, the more "
                "mentions are found.",
        "cols": ["name", "role", "agency", "aliases", "active", "notes"],
        "example": ["EXAMPLE — delete this row", "Resident Coordinator",
                    "UNRCO", "Ж. Смит, Smith Jane, J. Smith", "yes",
                    "How a row should look"],
        "rows": [
            ["<FILL IN>", "UN Resident Coordinator", "UNRCO", "", "yes",
             "Add the real name and every Cyrillic spelling"],
            ["<FILL IN>", "Head of Office", "UNICEF", "", "yes", ""],
            ["<FILL IN>", "Head of Office", "UNDP", "", "yes", ""],
            ["<FILL IN>", "Head of Office", "UNFPA", "", "yes", ""],
            ["<FILL IN>", "Representative", "WHO", "", "yes", ""],
            ["<FILL IN>", "Head of Office", "IOM", "", "yes", ""],
            ["<FILL IN>", "Representative", "FAO", "", "yes", ""],
            ["<FILL IN>", "Head of Office", "UNESCO", "", "no", ""],
            ["<FILL IN>", "Head of Office", "ILO", "", "no", ""],
            ["<FILL IN>", "Head of Office", "UN Women", "", "no", ""],
        ],
    },
    # --------------------------------------------------------------- un_sites
    "un_sites": {
        "help": "Our own publications. These are indexed so that when an "
                "outlet cites us, the report can link back to the exact press "
                "release or report being cited. feed_url is optional.",
        "cols": ["agency", "name", "url", "feed_url", "active", "notes"],
        "example": ["EXAMPLE — delete this row", "Example office",
                    "https://example.un.org/en/news", "", "yes",
                    "How a row should look"],
        "rows": [
            ["UNRCO", "UN Mongolia — stories",
             "https://mongolia.un.org/en/stories", "", "yes", ""],
            ["UNRCO", "UN Mongolia — press releases",
             "https://mongolia.un.org/en/resources/press-releases", "", "yes",
             "VERIFY path"],
            ["UNICEF", "UNICEF Mongolia — press",
             "https://www.unicef.org/mongolia/press-releases", "", "yes", ""],
            ["UNDP", "UNDP Mongolia — press",
             "https://www.undp.org/mongolia/press-releases", "", "yes", ""],
            ["UNFPA", "UNFPA Mongolia — news",
             "https://mongolia.unfpa.org/en/news", "", "yes", ""],
            ["WHO", "WHO Mongolia — news",
             "https://www.who.int/mongolia/news", "", "yes", ""],
            ["IOM", "IOM Mongolia", "https://mongolia.iom.int/news", "", "yes",
             "VERIFY path"],
            ["FAO", "FAO Mongolia", "https://www.fao.org/mongolia/news", "",
             "yes", "VERIFY path"],
            ["UNESCO", "UNESCO", "https://www.unesco.org/en/articles", "", "no",
             ""],
            ["ILO", "ILO", "https://www.ilo.org/resource/news", "", "no", ""],
            ["UN Women", "UN Women Asia-Pacific",
             "https://asiapacific.unwomen.org/en/news-and-events", "", "no", ""],
            ["ReliefWeb", "ReliefWeb Mongolia updates",
             "https://reliefweb.int/country/mng", "", "yes", ""],
        ],
    },
    # -------------------------------------------------------------- glossary
    "glossary": {
        "help": "Agreed Mongolian wording for recurring United Nations terms. "
                "Anything listed here is used verbatim in the Mongolian "
                "translation instead of being re-translated each week.",
        "cols": ["english", "mongolian", "notes"],
        "example": ["EXAMPLE — delete this row", "ЖИШЭЭ", "How a row should look"],
        "rows": [
            ["United Nations", "Нэгдсэн Үндэстний Байгууллага", ""],
            ["Resident Coordinator", "Суурин зохицуулагч", "confirm office usage"],
            ["Sustainable Development Goals", "Тогтвортой хөгжлийн зорилтууд", ""],
            ["United Nations Country Team", "НҮБ-ын Улсын багийн", "confirm"],
            ["Cooperation Framework", "Хамтын ажиллагааны хүрээ", "confirm"],
            ["humanitarian assistance", "хүмүүнлэгийн тусламж", ""],
            ["climate change", "уур амьсгалын өөрчлөлт", ""],
            ["dzud", "зуд", ""],
            ["herder households", "малчин өрх", ""],
            ["air pollution", "агаарын бохирдол", ""],
            ["social protection", "нийгмийн хамгаалал", ""],
            ["gender equality", "жендэрийн тэгш байдал", ""],
        ],
    },
}

WIDTHS = {"name": 34, "type": 15, "query_or_url": 52, "language": 11,
          "default_category": 32, "active": 9, "notes": 42, "domain": 22,
          "country": 15, "priority": 10, "rss_url": 40, "role": 26,
          "agency": 14, "aliases": 46, "url": 52, "feed_url": 34,
          "english": 34, "mongolian": 40}


def readme_sheet(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("Mongolia Weekly — shared source list", True),
        ("", False),
        ("This is the only file the office needs to edit. The Monday run reads it.", False),
        ("", False),
        ("HOW TO USE", True),
        ("1. Edit any sheet in Excel as usual. One row = one source.", False),
        ("2. Save the file, keeping the name sources.xlsx.", False),
        ("3. Upload it back to GitHub: open the repository, go to the sources", False),
        ("   folder, click 'Add file' > 'Upload files', drag the file in,", False),
        ("   then 'Commit changes'.", False),
        ("4. The change takes effect at the next Monday run.", False),
        ("", False),
        ("RULES", True),
        ("- Column 'active': write yes or no. A row set to no is ignored.", False),
        ("- Never rename a sheet or a column header; the scripts read them by name.", False),
        ("- Delete the grey EXAMPLE row in each sheet once you have understood it.", False),
        ("- A row that cannot be read does not stop the run: it is listed under", False),
        ("  'Source problems' in the Monday pull request so it can be fixed.", False),
        ("- Cells marked <FILL IN> must be completed before the sheet is useful.", False),
        ("", False),
        ("SHEETS", True),
        ("sources   — feeds and searches for the weekly news report", False),
        ("outlets   — media outlets watched for mentions of the United Nations", False),
        ("roster    — UN officials whose names may be quoted, plus their aliases", False),
        ("un_sites  — our own publications, indexed to link mentions back to source", False),
        ("glossary  — agreed Mongolian wording for recurring UN terms", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", size=11 if bold else 10, bold=bold)
    ws.column_dimensions["A"].width = 82
    ws.sheet_view.showGridLines = False
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)
    readme_sheet(wb)

    for name, spec in SHEETS.items():
        ws = wb.create_sheet(name)
        # help banner
        ws.cell(row=1, column=1, value=spec["help"])
        ws.cell(row=1, column=1).font = Font(name="Arial", size=9, italic=True)
        ws.cell(row=1, column=1).fill = NOTE_FILL
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(spec["cols"]))
        ws.row_dimensions[1].height = 46

        for j, col in enumerate(spec["cols"], 1):
            c = ws.cell(row=2, column=j, value=col)
            c.font, c.fill, c.border = HEAD_FONT, HEAD_FILL, THIN
            c.alignment = Alignment(vertical="center")
            ws.column_dimensions[c.column_letter].width = WIDTHS.get(col, 18)

        for j, val in enumerate(spec["example"], 1):
            c = ws.cell(row=3, column=j, value=val)
            c.font, c.border = EXAMPLE_FONT, THIN

        for i, row in enumerate(spec["rows"], 4):
            for j, val in enumerate(row, 1):
                c = ws.cell(row=i, column=j, value=val)
                c.font, c.border = BODY_FONT, THIN

        if "active" in spec["cols"]:
            col_letter = ws.cell(row=2, column=spec["cols"].index("active") + 1).column_letter
            dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}3:{col_letter}400")

        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)} with sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    if OUT.exists() and "--force" not in sys.argv:
        print(f"{OUT.relative_to(ROOT)} already exists. Pass --force to overwrite "
              f"(this destroys the office's edits).")
        sys.exit(1)
    build()
